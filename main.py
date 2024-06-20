import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class Annotation:
    def __init__(self, csv_file_path=None, images_folder=None):
        self.csv_file_path = csv_file_path
        self.images_folder = images_folder
        self.df = None
        if csv_file_path:
            self.df = pd.read_csv(csv_file_path, delimiter='\t')  # Specify tab delimiter
        self.out_put_field = 'clean'

    def merge_excel_cells(self, excel_file_path, output_file_path):
        # Load the workbook and select the sheet
        workbook = load_workbook(filename=excel_file_path)
        sheet = workbook['Sheet1']

        def merge_cells(sheet, start_row, end_row, column):
            # Get the column letter
            col_letter = get_column_letter(column)
            # Merge the cells
            sheet.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')

        # Loop through columns and merge cells with the same values
        for col in range(1, sheet.max_column + 1):
            start_row = 1
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=col).value != sheet.cell(row=start_row, column=col).value:
                    if row - 1 > start_row:
                        merge_cells(sheet, start_row, row - 1, col)
                    start_row = row
            # Handle the last set of cells
            if sheet.cell(row=sheet.max_row, column=col).value == sheet.cell(row=start_row, column=col).value:
                merge_cells(sheet, start_row, sheet.max_row, col)

        # Save the modified workbook
        workbook.save(output_file_path)
        return output_file_path

# Example usage
annotation = Annotation()
merged_file_path = annotation.merge_excel_cells(excel_file_path='New_BOOK.xlsx', output_file_path='Merged_BOOK.xlsx')
print(f'Merged Excel file saved at: {merged_file_path}')
