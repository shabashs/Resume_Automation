import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

class Annotation:
    def __init__(self, csv_file_path=None, images_folder=None):
        self.csv_file_path = csv_file_path
        self.images_folder = images_folder
        self.df = None
        if csv_file_path:
            self.df = pd.read_csv(csv_file_path, delimiter='\t')  # Specify tab delimiter
        self.out_put_field = 'clean'

    def split_and_merge(self, input_excel_path, output_file_path='Merged_BOOK1.xlsx'):
        # Load the data into a DataFrame
        df = pd.read_excel(input_excel_path)

        # Split values
        new_rows = []
        for _, row in df.iterrows():
            row = row.astype(str)
            max_splits = max([str(cell).count('|') for cell in row]) + 1
            split_data = [str(cell).split('|') if '|' in str(cell) else [cell] * max_splits for cell in row]
            for i in range(max_splits):
                new_row = [split_data[col][i].strip() if i < len(split_data[col]) else '' for col in range(len(row))]
                new_rows.append(new_row)

        # Create a new DataFrame with the split data
        new_df = pd.DataFrame(new_rows, columns=df.columns)

        # Create a new workbook and add the split data
        workbook = Workbook()
        sheet = workbook.active
        for r in dataframe_to_rows(new_df, index=False, header=True):
            sheet.append(r)

        # Merge cells in the new sheet
        def merge_cells(sheet, start_row, end_row, column):
            col_letter = get_column_letter(column)
            sheet.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')

        for col in range(1, sheet.max_column + 1):
            start_row = 1
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=col).value != sheet.cell(row=start_row, column=col).value:
                    if row - 1 > start_row:
                        merge_cells(sheet, start_row, row - 1, col)
                    start_row = row
            if sheet.cell(row=sheet.max_row, column=col).value == sheet.cell(row=start_row, column=col).value:
                merge_cells(sheet, start_row, sheet.max_row, col)

        # Save the modified workbook
        workbook.save(output_file_path)
        print("Data has been split and merged. Saved to:", output_file_path)
        return output_file_path

# Example usage
annotation = Annotation()
merged_file_path = annotation.split_and_merge(input_excel_path='extracted_data.xlsx', output_file_path='Merged_BOOK1.xlsx')
print(f'Merged Excel file saved at: {merged_file_path}')
